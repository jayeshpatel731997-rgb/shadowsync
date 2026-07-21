"""Create the fixed, synthetic Phase 1 ShadowSync demonstration assets."""

from __future__ import annotations

import argparse
import csv
import sqlite3
from pathlib import Path
from typing import Final

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

FIXED_SEED: Final[int] = 20260721
FIXED_UPDATED_AT: Final[str] = "2026-07-21T14:00:00Z"


def _records() -> dict[str, list[dict[str, object]]]:
    """Return deterministic system-of-record rows; no personal or proprietary data."""
    return {
        "open_orders": [
            {"order_id": "PO-1001", "sku": "SKU-A100", "vendor_id": "V-001", "quantity_each": 120, "unit_price_cents": 1250, "status": "OPEN", "expected_date": "2026-08-01", "updated_at": FIXED_UPDATED_AT, "row_version": 1},
            {"order_id": "PO-1002", "sku": "SKU-B200", "vendor_id": "V-002", "quantity_each": 48, "unit_price_cents": 875, "status": "OPEN", "expected_date": "2026-08-04", "updated_at": FIXED_UPDATED_AT, "row_version": 1},
            {"order_id": "PO-1003", "sku": "SKU-C300", "vendor_id": "V-001", "quantity_each": 72, "unit_price_cents": 2100, "status": "HOLD", "expected_date": "2026-08-07", "updated_at": FIXED_UPDATED_AT, "row_version": 1},
        ],
        "inventory": [
            {"sku": "SKU-A100", "on_hand_each": 240, "uom": "EA", "warehouse_zone": "A-01", "updated_at": FIXED_UPDATED_AT, "row_version": 1},
            {"sku": "SKU-B200", "on_hand_each": 96, "uom": "EA", "warehouse_zone": "B-02", "updated_at": FIXED_UPDATED_AT, "row_version": 1},
            {"sku": "SKU-C300", "on_hand_each": 30, "uom": "EA", "warehouse_zone": "C-03", "updated_at": FIXED_UPDATED_AT, "row_version": 1},
        ],
        "allocations": [
            {"allocation_id": "AL-001", "order_id": "PO-1001", "sku": "SKU-A100", "allocated_each": 40, "updated_at": FIXED_UPDATED_AT, "row_version": 1},
            {"allocation_id": "AL-002", "order_id": "PO-1002", "sku": "SKU-B200", "allocated_each": 20, "updated_at": FIXED_UPDATED_AT, "row_version": 1},
        ],
        "vendor_terms": [
            {"vendor_id": "V-001", "vendor_name": "Synthetic Supply Co", "payment_terms": "NET30", "lead_time_days": 10, "contact_email": "ops@example.invalid", "updated_at": FIXED_UPDATED_AT, "row_version": 1},
            {"vendor_id": "V-002", "vendor_name": "Demo Components Ltd", "payment_terms": "NET45", "lead_time_days": 14, "contact_email": "planning@example.invalid", "updated_at": FIXED_UPDATED_AT, "row_version": 1},
        ],
    }


def connect_database(path: Path) -> sqlite3.Connection:
    """Open SQLite with safety settings that must be enabled per connection."""
    connection = sqlite3.connect(path)
    connection.execute("PRAGMA foreign_keys = ON")
    return connection


def ensure_database_schema(path: Path) -> None:
    """Apply idempotent schema additions to an existing SQLite database."""
    schema = Path(__file__).with_name("schema.sql").read_text(encoding="utf-8")
    with connect_database(path) as connection:
        connection.executescript(schema)


def create_database(path: Path) -> None:
    """Create a fresh SQLite system of record at *path*."""
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        path.unlink()
    ensure_database_schema(path)
    with connect_database(path) as connection:
        for table, rows in _records().items():
            columns = list(rows[0])
            placeholders = ", ".join("?" for _ in columns)
            connection.executemany(
                f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({placeholders})",
                ([row[column] for column in columns] for row in rows),
            )


def create_excel(path: Path) -> None:
    """Create an analyst workbook with intentional, documented drift."""
    records = _records()
    orders = pd.DataFrame(records["open_orders"])
    orders.loc[orders.order_id == "PO-1001", "unit_price_cents"] = 1195  # stale_price
    orders.loc[orders.order_id == "PO-1002", "quantity_each"] = 60  # quantity_mismatch

    inventory = pd.DataFrame(records["inventory"])
    inventory.loc[inventory.sku == "SKU-A100", "uom"] = "Each"  # uom_inconsistency

    allocations = pd.DataFrame(records["allocations"])
    orphan = allocations.iloc[0].copy()
    orphan["allocation_id"] = "AL-999"
    orphan["order_id"] = "PO-MISSING"
    allocations = pd.concat([allocations, orphan.to_frame().T], ignore_index=True)

    terms = pd.DataFrame(records["vendor_terms"])
    terms.loc[terms.vendor_id == "V-002", "contact_email"] = None  # missing_attribute

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        orders.to_excel(writer, sheet_name="open_orders", index=False)
        inventory.to_excel(writer, sheet_name="inventory", index=False)
        allocations.to_excel(writer, sheet_name="allocations", index=False)
        terms.to_excel(writer, sheet_name="vendor_terms", index=False)
        pd.DataFrame(
            [
                ("fixed_seed", FIXED_SEED),
                ("data_classification", "SYNTHETIC"),
                ("purpose", "learning/demo only"),
                ("intentional_drift", "stale price; orphaned allocation; quantity mismatch; UOM inconsistency; missing attribute"),
            ],
            columns=("property", "value"),
        ).to_excel(writer, sheet_name="README", index=False)

    # Keep the generated workbook usable for an analyst while preserving typed cells.
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for column in worksheet.columns:
            values = ["" if cell.value is None else str(cell.value) for cell in column]
            worksheet.column_dimensions[column[0].column_letter].width = min(max(len(value) for value in values) + 2, 48)
    workbook.save(path)


def create_bi_extract(path: Path) -> None:
    """Create a flattened simulated Power BI dataset with deliberate drift."""
    orders = pd.DataFrame(_records()["open_orders"])
    orders.loc[orders.order_id == "PO-1003", "quantity_each"] = 70
    orders.loc[orders.order_id == "PO-1001", "unit_price_cents"] = 1195
    orders.insert(0, "dataset", "open_orders")
    orders.to_csv(path, index=False, lineterminator="\n")


def create_manual_baseline(path: Path) -> None:
    """Create a synthetic historical baseline for later KPI comparison."""
    rows = [
        ("MB-001", "stale_price", 95, "manual_research"),
        ("MB-002", "orphaned_allocation", 140, "manual_research"),
        ("MB-003", "quantity_mismatch", 75, "manual_update"),
        ("MB-004", "uom_inconsistency", 55, "manual_update"),
        ("MB-005", "missing_attribute", 45, "manual_update"),
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(("case_id", "drift_type", "minutes_to_reconcile", "resolution_method"))
        writer.writerows(rows)


def generate_demo(output_dir: Path) -> dict[str, Path]:
    """Generate all Phase 1 assets and return their paths."""
    output_dir.mkdir(parents=True, exist_ok=True)
    assets = {
        "database": output_dir / "shadowsync.db",
        "excel": output_dir / "analyst_shadow.xlsx",
        "bi_extract": output_dir / "power_bi_extract.csv",
        "manual_baseline": output_dir / "manual_baseline.csv",
    }
    create_database(assets["database"])
    create_excel(assets["excel"])
    create_bi_extract(assets["bi_extract"])
    create_manual_baseline(assets["manual_baseline"])
    return assets


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", type=Path, default=Path("demo"))
    args = parser.parse_args()
    for name, path in generate_demo(args.output_dir).items():
        print(f"{name}: {path}")


if __name__ == "__main__":
    main()
